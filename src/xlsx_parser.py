"""
Python 3.8
Author: Adam Turner <turner.adch@gmail.com>
"""

# python package index
from openpyxl import load_workbook
# standard library
import datetime


class XlsxTable(object):

    def __init__(self):
        self.config = None

    def define_config(self):
        pass


class XlsxParser(object):

    def __init__(self):
        pass

    @staticmethod
    def extract_cell_data(sheet, rows, cols):
        """Extracts a single data element from a defined cell area.

        Args:
            sheet: openpyxl Worksheet
            rows: list[int], list of rows to extract from, example: [1, 2, 3]
            cols: list[str], list of columns to extra from, example: ['A', 'B']

        Returns:
            extract: str, concatenated cell data, example: A1 + B1 + A2 + B2 + A3 + B3
        """
        extract = []
        for row in rows:
            text = []
            for col in cols:
                cell_name = col + str(row)
                cell_value = str(sheet[cell_name].value).strip()
                cell_value = sheet[cell_name].value
                
                if isinstance(cell_value, datetime.datetime):
                    cell_value = cell_value.strftime("%Y-%m-%d")
                elif cell_value is None:
                    continue
                else:
                    cell_value = str(cell_value).strip()
                
                text.append(cell_value)
            text = " ".join(text)
            if text:
                extract.append(text)
            else:
                continue
        if extract:
            return " ".join(extract)
        else:
            return ""

    def define_tables(self):
        self.table_configs = [
            {
                'cols': ['dock_arrival', 'flag', 'last_port', 'loa', 'eta', 'etd', 'dock', 'service', 'agent'],
                'groups': ['A', 'B', ('C', 'D'), 'E', 'F', 'G', 'H', 'I', 'J']
            },
            {
                'cols': ['arrival', 'barge', 'tug', 'dock', 'service', 'agent'],
                'groups': ['A', ('B', 'C', 'D'), ('E', 'F', 'G'), 'H', 'I', 'J']
            }
        ]
        # validate and fill out configs
        for config in table_configs:
            if len(config['cols']) != len(config['groups']):
                raise ValueError(f"{config} cols/groups length mismatch!")
            else:
                standard_config = {'start': None, 'end': None, 'data': None, 'header': 3, 'row_depth': 2}
                for key in standard_config.keys():
                    if key not in config:
                        config[key] = standard_config[key]
        # define table names
        # TODO: consider making a table object that the parser can parse
        tables = {
            'in port': table_configs[0].copy(),
            'anchorage': table_configs[0].copy(),
            'vessels due': table_configs[0].copy(),
            'amfels': table_configs[0].copy(),
            'barges in port': table_configs[1].copy()
        }
        table_names = list(tables.keys())

    def main(self, file_path):
        wb = load_workbook(file_path)
        sheet = wb['Vessel and Barge Arrival Sheet']
        # TODO: implement something that looks for sheet names
        # and matches against a regular expression for validation
        # wb.sheetnames returns list[str]
        # sheet_list = wb.sheetnames
        
        # loop through all cells in column 'A' trying to match
        # table names from the supplied table configs
        # figure out on which row each one starts and ends

        # table cache remembers the last table name that
        # was parsed until we parse a 'notes' cell
        # at which point it resets to empty str

        table_cache = ''
        for row in range(sheet.min_row, sheet.max_row):
            cell_name = f"A{row}"
            # encountered rogue datetime.datetime during testing, so cast to str
            cell_value = str(sheet[cell_name].value).strip().lower()

            if cell_value in table_names:
                # cell value can now be considered table name
                if table_cache == '':
                    tables[cell_value]['start'] = row
                    table_cache = cell_value
                    print(f"Detected '{cell_value}' table starts on row {row}")
                else:
                    raise ValueError(f"Parsed '{cell_value}' table before parsing 'Notes:' for '{table_cache}' table.")        

            elif cell_value == 'notes:':
                if table_cache == '':
                    raise ValueError(f"On cell {cell_name}, parsed 'Notes:' while table_cache is empty.")
                else:
                    tables[table_cache]['end'] = row
                    print(f"  > '{table_cache}' table ends on line {row}")
                    table_cache = ''
        print(tables)

        # TODO: consider integrating this into a single loop through the data
        for table_name in tables.keys():
            table = tables[table_name]
            records = []
            for row in range(table['start']+table['header'], table['end'], table['row_depth']):
                vector = []
                for group in table['groups']:
                    row_end = row + table['row_depth'] - 1
                    data = extract_cell_data(sheet, rows=[row, row_end], cols=group)
                    vector.append(data)
                    continue
                records.append(tuple(vector))
                continue
            df = pd.DataFrame.from_records(records)
            df.columns = table['cols']
            df = df.replace(r"^\s?$", None, regex=True)
            table['data'] = df
            breakpoint()


        inport = tables['in port']
        df = inport['data']
        breakpoint()
        return None
